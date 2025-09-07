import os
import tempfile
from django.test import TestCase
from unittest.mock import MagicMock, patch
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.exceptions import ParseError
from unittest.mock import MagicMock
from .views import UploadView
from .helpers.gestor_excel_sheet_data import GestorExcelSheetData
from .helpers.gestor_excel_file_data import GestorExcelFileData
from .helpers.columns_configuration import DataTypeCell, PatientData


class GestorExcelSheetDataTest(TestCase):
  """
  Clase de pruebas unitarias para GestorExcelSheetData.
  """

  def setUp(self):
    """
    Configura el entorno de prueba. Se ejecuta antes de cada método de prueba.
    """
    self.gestor = GestorExcelSheetData()
    self.workbook = Workbook()
    self.sheet: Worksheet = self.workbook.active
    self.sheet.title = "Trauma Register"
    
    # Llenamos la hoja de cálculo con datos de prueba
    self.sheet['A1'] = 'Header1'
    self.sheet['B1'] = 'Header2'
    self.sheet['A2'] = 'Patient1'
    self.sheet['B2'] = 100
    self.sheet['A3'] = 'Patient2'
    self.sheet['B3'] = 200
    self.sheet['A4'] = None # Celda vacía para prueba de define_limit_row

  def test_cell_get_value_with_str_converts_to_string(self):
    """
    Prueba que el método convierte el valor de una celda a string correctamente.
    """
    # Prueba con un valor numérico
    result = self.gestor.cell_get_value_with_str(cell_name="B2", sheet=self.sheet)
    self.assertEqual(result, "100")
    
    # Prueba con un valor de texto
    result_str = self.gestor.cell_get_value_with_str(cell_name="A2", sheet=self.sheet)
    self.assertEqual(result_str, "Patient1")
      
  def test_cell_get_value_with_str_returns_original_value(self):
    """
    Prueba que el método retorna el valor original sin convertir si 'convert_str_answer' es False.
    """
    result = self.gestor.cell_get_value_with_str(cell_name="B2", sheet=self.sheet, convert_str_answer=False)
    self.assertEqual(result, 100)
      
  def test_get_row_elements(self):
    """
    Prueba que el método get_row_elements extrae los datos de una fila correctamente.
    """
    column_ubication = {"patient_id": "A", "age": "B"}
    row_number = 2
    
    expected_result = {"patient_id": "Patient1", "age": "100"}
    result = self.gestor.get_row_elements(
      column_ubication=column_ubication, 
      sheet=self.sheet, 
      row_number=row_number
    )
    self.assertEqual(result, expected_result)
      
  def test_define_limit_row(self):
    """
    Prueba que el método define_limit_row encuentra la primera fila vacía.
    """
    # La celda A4 está vacía en nuestro setUp
    expected_limit_row = 4 
    result = self.gestor.define_limit_row(sheet=self.sheet, selected_test_column="A")
    self.assertEqual(result, expected_limit_row)

  def test_get_data_from_sheet(self):
    """
    Prueba la función principal get_data_from_sheet.
    """
    # Mockeamos `define_limit_row` para evitar la lógica interna en esta prueba
    # Así probamos que `get_data_from_sheet` usa el valor del límite correctamente.
    self.gestor.define_limit_row = MagicMock(return_value=4)
    
    column_ubication = {"patient_id": "A", "age": "B"}
    
    expected_result = [
      {"patient_id": "Patient1", "age": "100"},
      {"patient_id": "Patient2", "age": "200"},
    ]
    
    result = self.gestor.get_data_from_sheet(
      column_ubication=column_ubication, 
      sheet=self.sheet
    )
    self.assertEqual(result, expected_result)

  def tearDown(self):
    """
    Limpia los recursos después de cada prueba.
    """
    self.workbook.close()

class GestorExcelFileDataTest(TestCase):
  """
  Clase de pruebas unitarias para GestorExcelFileData.
  """

  def setUp(self):
    """
    Configura el entorno de prueba.
    Crea un archivo de Excel temporal para simular la lectura.
    """
    # Configuramos los datos de prueba
    self.sheets_config = {
      "Hoja1:Datos": {"col1": "A", "col2": "B"},
      "Hoja2:Info": {"col3": "C", "col4": "D"},
    }

    # Creamos un archivo de Excel temporal para simular la lectura
    self.temp_dir = tempfile.mkdtemp()
    self.file_path = os.path.join(self.temp_dir, "test_file.xlsx")

    # Creamos el archivo de Excel con los nombres de hoja esperados
    workbook = Workbook()
    workbook.create_sheet("Hoja1")
    workbook.create_sheet("Hoja2")
    workbook.save(self.file_path)
    workbook.close()

  def tearDown(self):
    """
    Limpia los recursos temporales creados.
    """
    os.remove(self.file_path)
    os.rmdir(self.temp_dir)

  # --- Pruebas para el método inform_bad_format ---
  
  @patch('openpyxl.load_workbook')
  def test_inform_bad_format_raises_error_for_missing_sheet(self, mock_load_workbook):
    """
    Prueba que inform_bad_format lanza un ParseError si falta una hoja.
    """
    # Configurar un mock para el libro de trabajo de openpyxl que no tiene una hoja.
    mock_workbook = MagicMock()
    mock_workbook.sheetnames = ["Hoja1"]  # Falta Hoja2
    mock_load_workbook.return_value = mock_workbook
    
    # Creamos una instancia con un archivo de Excel falso
    gestor = GestorExcelFileData(
      sheets_columns_configuration=self.sheets_config,
      gestor_sheet=MagicMock(),
      file_xls=self.file_path
    )

    with self.assertRaises(ParseError) as context:
      gestor.inform_bad_format()

    expected_message = 'No se pudo realizar la carga debido a que el formato no es correcto por la falta de la hoja "Hoja2" en el archivo Excel.\nPor favor utilice un formato válido.'
    self.assertEqual(str(context.exception), expected_message)
      
  @patch('openpyxl.load_workbook')
  def test_inform_bad_format_passes_with_correct_sheets(self, mock_load_workbook):
    """
    Prueba que inform_bad_format no lanza un error si las hojas existen.
    """
    mock_workbook = MagicMock()
    mock_workbook.sheetnames = ["Hoja1", "Hoja2"]
    mock_load_workbook.return_value = mock_workbook

    gestor = GestorExcelFileData(
      sheets_columns_configuration=self.sheets_config,
      gestor_sheet=MagicMock(),
      file_xls=self.file_path
    )
    
    # No debería levantar ninguna excepción
    try:
      gestor.inform_bad_format()
    except ParseError:
      self.fail("inform_bad_format levantó un ParseError inesperadamente")

  # --- Pruebas para el método get_data_from_file ---

  @patch('openpyxl.load_workbook')
  def test_get_data_from_file_extracts_data_correctly(self, mock_load_workbook):
    """
    Prueba que get_data_from_file extrae datos de todas las hojas correctamente.
    """
    # 1. Mockeamos la dependencia de GestorExcelFileData
    # Simulamos una instancia de GestorExcelSheetData
    mock_gestor_sheet = MagicMock()

    # Configuramos lo que get_data_from_sheet debe devolver para cada llamada
    mock_gestor_sheet.get_data_from_sheet.side_effect = [
      [{"row_data_sheet1": "data1"}],
      [{"row_data_sheet2": "data2"}],
    ]

    # 2. Simulamos el libro de trabajo de openpyxl
    # Creamos mocks para las hojas
    mock_sheet1 = MagicMock()
    mock_sheet2 = MagicMock()
    
    # Configuramos el mock del libro de trabajo para que devuelva los mocks de hoja
    mock_workbook = MagicMock()
    mock_workbook.__getitem__.side_effect = {
      "Hoja1": mock_sheet1,
      "Hoja2": mock_sheet2,
    }.get
    
    # Le decimos al patch qué devolver cuando se llame a load_workbook
    mock_load_workbook.return_value = mock_workbook

    # 3. Creamos la instancia del objeto a probar
    gestor = GestorExcelFileData(
      sheets_columns_configuration=self.sheets_config,
      gestor_sheet=mock_gestor_sheet,
      file_xls=self.file_path
    )

    # 4. Ejecutamos la prueba
    result = gestor.get_data_from_file()

    # 5. Verificamos que los datos devueltos sean correctos
    expected_result = {
      "Hoja1:Datos": [{"row_data_sheet1": "data1"}],
      "Hoja2:Info": [{"row_data_sheet2": "data2"}],
    }
    self.assertEqual(result, expected_result)
    
    # 6. Verificamos que los métodos del mock fueron llamados correctamente
    mock_gestor_sheet.get_data_from_sheet.assert_any_call(
      column_ubication={"col1": "A", "col2": "B"}, sheet=mock_sheet1
    )
    mock_gestor_sheet.get_data_from_sheet.assert_any_call(
      column_ubication={"col3": "C", "col4": "D"}, sheet=mock_sheet2
    )
    
    # Es una buena práctica verificar también el número total de llamadas
    self.assertEqual(mock_gestor_sheet.get_data_from_sheet.call_count, 2)
    
    # Y que el libro de trabajo se cerró correctamente
    mock_workbook.close.assert_called_once()

  def test_get_data_from_file_handles_exception_gracefully(self):
    """
    Prueba que get_data_from_file maneja excepciones internas sin propagarlas.
    """
    # Configuramos un mock para `get_data_from_sheet` que lanza una excepción
    mock_gestor_sheet = MagicMock(spec=GestorExcelSheetData)
    mock_gestor_sheet.get_data_from_sheet.side_effect = Exception("Simulated read error")
    
    # No es necesario mockear `openpyxl.load_workbook` porque el error ocurre después.
    gestor = GestorExcelFileData(
      sheets_columns_configuration=self.sheets_config,
      gestor_sheet=mock_gestor_sheet,
      file_xls=self.file_path
    )
    
    # La función debería capturar la excepción y devolver None
    # O dependiendo de la lógica, podría devolver un diccionario vacío.
    # En tu código actual, `get_data_from_file` devuelve `None` en caso de excepción.
    result = gestor.get_data_from_file()
    self.assertIsNone(result)
  

class UploadViewTest(TestCase):
  """
  Clase de pruebas unitarias para la vista UploadView.
  """

  def setUp(self):
    # Instancia de la vista para probar sus métodos auxiliares
    self.view = UploadView()

    # Mocks para dependencias
    self.mock_patient_model = MagicMock()
    self.mock_model_1 = MagicMock()
    self.mock_model_2 = MagicMock()
    
    # Diccionarios de prueba
    self.patient_data_list = [
      {"trauma_register_record_id": "1", "name": "Patient 1"},
      {"trauma_register_record_id": "2", "name": "Patient 2"},
      {"trauma_register_record_id": "3", "name": "Patient 3"},
    ]
    
    self.data_file_for_save = {
      "PatientData:1": [
        {"trauma_register_record_id": "1", "string_field": "test string", "int_field": "10", "bool_field": "si", "date_field": "10/01/2023"},
      ],
      "OtherData:2": [
        {"trauma_register_record_id": "1", "int_field": "120", "minute_field": "15"},
      ]
    }
    
    self.columns_config_for_save = {
      "PatientData:1": {
        "model": self.mock_patient_model,
        "type": {
          "trauma_register_record_id": DataTypeCell.INT,
          "string_field": DataTypeCell.STRING,
          "int_field": DataTypeCell.INT,
          "bool_field": DataTypeCell.BOOLEAN,
          "date_field": DataTypeCell.DATE,
        }
      },
      "OtherData:2": {
        "model": self.mock_model_1,
        "type": {
          "trauma_register_record_id": DataTypeCell.INT,
          "int_field": DataTypeCell.INT,
          "minute_field": DataTypeCell.INT,
        }
      }
    }
      
  # --- Pruebas para el método delete_existing_data ---

  @patch('upload_manager.views.PatientData.objects')
  def test_delete_existing_data_deletes_patients(self, mock_objects):
    """
    Prueba que el método delete_existing_data elimina pacientes existentes y devuelve sus IDs.
    """
    # Configuramos el mock para que `get` devuelva un objeto, simulando que el paciente existe.
    mock_patient_1 = MagicMock()
    mock_patient_1.delete.return_value = None
    
    # Simulamos que el paciente 1 existe y el 2 y 3 no
    def mock_get(trauma_register_record_id):
      if trauma_register_record_id == 1:
        return mock_patient_1
      raise PatientData.DoesNotExist
    
    mock_objects.get.side_effect = mock_get
    
    # Ejecutamos el método
    patient_data = [{"trauma_register_record_id": "1"}]
    repeteated_users = self.view.delete_existing_data(patient_data=patient_data)
    
    # Verificamos que el método get y delete fueron llamados
    mock_objects.get.assert_called_with(trauma_register_record_id=1)
    mock_patient_1.delete.assert_called_once()
    
    # Verificamos el resultado
    self.assertEqual(repeteated_users, ["1"])

  @patch('upload_manager.views.PatientData.objects')
  def test_delete_existing_data_skips_non_existing_patients(self, mock_objects):
    """
    Prueba que el método delete_existing_data no hace nada para pacientes que no existen.
    """
    # Configuramos el mock para que `get` siempre lance DoesNotExist
    mock_objects.get.side_effect = PatientData.DoesNotExist
    
    patient_data = [{"trauma_register_record_id": "1"}]
    repeteated_users = self.view.delete_existing_data(patient_data=patient_data)

    # Verificamos que se llamó a get, pero delete no
    mock_objects.get.assert_called_with(trauma_register_record_id=1)
    self.assertEqual(repeteated_users, [])

  # --- Pruebas para el método search_existing_data ---

  def test_search_existing_data_finds_existing(self):
    """
    Prueba que search_existing_data devuelve True si el dato existe.
    """
    # Creamos un mock de un modelo que tiene un gestor 'objects'
    mock_model = MagicMock()
    mock_model.objects.filter.return_value.exists.return_value = True
    
    result = self.view.search_existing_data(model=mock_model, trauma_register_record_id=1)
    
    self.assertTrue(result)
    mock_model.objects.filter.assert_called_once_with(trauma_register_record_id=1)

  def test_search_existing_data_finds_none(self):
    """
    Prueba que search_existing_data devuelve False si el dato no existe.
    """
    mock_model = MagicMock()
    mock_model.objects.filter.return_value.exists.return_value = False
    
    result = self.view.search_existing_data(model=mock_model, trauma_register_record_id=1)
    
    self.assertFalse(result)
    mock_model.objects.filter.assert_called_once_with(trauma_register_record_id=1)

  # --- Pruebas para el método save_elements_by_model ---
  
  @patch('upload_manager.views.UploadView.search_existing_data', return_value=False)
  def test_save_elements_by_model_saves_new_data(self, mock_search):
    """
    Prueba que el método guarda datos nuevos cuando 'only_update' es False.
    """
    with patch('upload_manager.views.PatientData.objects.get') as mock_get:
      mock_get.return_value = MagicMock(trauma_register_record_id=1)
      
      # Mockeamos el método save del modelo
      self.mock_patient_model.return_value = MagicMock()
      self.mock_model_1.return_value = MagicMock()
      
      result = self.view.save_elements_by_model(
        data_file=self.data_file_for_save,
        column_name_type_to_model=self.columns_config_for_save,
        only_update=False,
      )
      
      # Verificamos que se intentó guardar los modelos
      self.mock_patient_model.assert_called_once()
      self.mock_model_1.assert_called_once()
      
      # Verificamos que se llamaron a los métodos save()
      self.mock_patient_model.return_value.save.assert_called_once()
      self.mock_model_1.return_value.save.assert_called_once()
      
      # Verificamos que no se encontraron errores
      self.assertEqual(result, [])

  @patch('upload_manager.views.UploadView.search_existing_data', return_value=True)
  def test_save_elements_by_model_skips_existing_when_only_create(self, mock_search):
    """
    Prueba que el método omite la creación de datos existentes cuando `existing_patients` es None.
    """
    with patch('upload_manager.views.PatientData.objects.get') as mock_get:
      mock_get.return_value = MagicMock(trauma_register_record_id=1)
      
      # Configuramos los mocks de los modelos para que lancen un error si se llama a save,
      # lo cual confirmará que no se llamaron.
      self.mock_patient_model.return_value.save.side_effect = AssertionError("Save should not be called")
      self.mock_model_1.return_value.save.side_effect = AssertionError("Save should not be called")

      result = self.view.save_elements_by_model(
        data_file=self.data_file_for_save,
        column_name_type_to_model=self.columns_config_for_save,
        only_update=False,
      )

      # Verificamos que no se encontraron errores
      self.assertEqual(result, [])
      # Las aserciones de save en los mocks de modelo validarán que no se llamó a save()
